# ==========================================================
# BOT KUNJUNGAN ANGGOTA (SINGLE FILE VERSION)
# Dengan fitur:
# - Import Excel
# - Foto kunjungan MULTI ID (HYBRID)
# - History pembayaran
# - Anti duplikasi foto (hash)
# - Hapus Kunjungan
# - /history ID
# - /rekap
# - Export Excel list belum dikunjungi
# ==========================================================

import os
import json
import hashlib
import datetime
from telegram import Update
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext
import openpyxl

# ==========================================================
# KONFIGURASI BOT
# ==========================================================

TOKEN = os.getenv("BOT_TOKEN")   # GANTI DENGAN TOKEN BOTFATHER

DATA_FILE = "data_kunjungan.json"
FOLDER_FOTO = "foto_kunjungan"
EXPORT_FOLDER = "exports"

os.makedirs(FOLDER_FOTO, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)

# ==========================================================
# LOAD & SAVE DATABASE JSON
# ==========================================================

def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

# ==========================================================
# FUNCTION: HASH FOTO UNTUK DETEKSI DUPLIKAT
# ==========================================================

def generate_photo_hash(file_path):
    """Menghasilkan hash unik dari file foto."""
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        buf = f.read()
        hasher.update(buf)
    return hasher.hexdigest()

# ==========================================================
# FUNCTION: MENAMBAHKAN STRUKTUR HISTORY KE ANGGOTA
# ==========================================================

def ensure_history_structure(data):
    """Pastikan setiap anggota memiliki struktur history."""
    changed = False
    for anggota in data:
        if "History" not in anggota:
            anggota["History"] = []
            changed = True
    if changed:
        save_data(data)
# ==========================================================
# BAGIAN 2 – IMPORT EXCEL & PARSING CAPTION
# ==========================================================

def import_excel(update: Update, context: CallbackContext):
    message = update.message
    doc = message.document

    if not doc:
        message.reply_text("Silakan upload file Excel (.xlsx).")
        return

    # Download file excel
    file_path = doc.get_file().download(custom_path="temp_excel.xlsx")

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        data = load_data()
        ensure_history_structure(data)

        next_no = len(data) + 1
        imported = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            ctr = row[0]
            id_anggota = row[1]
            nama = row[2]
            minggon = row[3]
            staff_raw = row[4]

            if not id_anggota or not nama:
                continue

            # Cegah duplikasi ID
            if any(d["ID"] == str(id_anggota) for d in data):
                continue

            # Bersihkan nama staff
            if staff_raw and "-" in str(staff_raw):
                staff_clean = str(staff_raw).split("-", 1)[1].strip()
            else:
                staff_clean = str(staff_raw).strip()

            anggota = {
                "No": next_no,
                "Ctr": str(ctr),
                "ID": str(id_anggota),
                "Nama": str(nama),
                "Minggon": str(minggon),
                "NamaStaff": staff_clean,
                "Status": "Belum Dikunjungi",
                "Foto": None,
                "TanggalKunjungan": None,
                "History": []
            }

            data.append(anggota)
            next_no += 1
            imported += 1

        save_data(data)
        message.reply_text(f"✔ Berhasil import {imported} anggota dari Excel.")

    except Exception as e:
        message.reply_text(f"❌ Gagal membaca Excel: {e}")

    finally:
        try:
            os.remove(file_path)
        except:
            pass


# ==========================================================
# PARSING CAPTION – HYBRID (DETEKSI TANGGAL OPSIONAL)
# ==========================================================

def parse_caption_hybrid(caption):
    """
    Mengembalikan:
    tanggal, list_baris
    """

    lines = caption.strip().split("\n")

    # Cek apakah baris pertama adalah tanggal
    first = lines[0].strip()
    tanggal_manual = None

    try:
        datetime.datetime.strptime(first, "%d-%m-%Y")
        tanggal_manual = first
        id_lines = lines[1:]  # sisanya baris ID
    except:
        # Tidak ada tanggal → pakai tanggal hari ini
        tanggal_manual = datetime.datetime.now().strftime("%d-%m-%Y")
        id_lines = lines[:]  # semua baris = ID

    return tanggal_manual, id_lines


# ==========================================================
# FUNCTION PARSE SATU BARIS "ID nominal"
# ==========================================================

def parse_id_payment_line(line):
    """
    Mengubah "ID nominal" → ("ID", nominal)
    """
    parts = line.strip().split()
    if len(parts) != 2:
        return None, None

    id_text = parts[0]

    try:
        payment = int(parts[1])
    except:
        return id_text, None

    return id_text, payment
# ==========================================================
# BAGIAN 3 – PROSES FOTO KUNJUNGAN (HYBRID + MULTI ID)
# ==========================================================

def foto_kunjungan(update: Update, context: CallbackContext):
    message = update.message

    if not message.photo:
        return

    caption = (message.caption or "").strip()

    if not caption:
        message.reply_text(
            "Format salah.\nGunakan:\n"
            "dd-mm-yyyy (opsional)\n"
            "ID nominal\n"
            "ID nominal"
        )
        return

    # ------------------------------------------------------
    # 1. LOAD DATABASE
    # ------------------------------------------------------
    data = load_data()
    ensure_history_structure(data)

    # ------------------------------------------------------
    # 2. PARSE TANGGAL DAN LIST BARIS ID
    # ------------------------------------------------------
    tanggal_kunjungan, id_lines = parse_caption_hybrid(caption)

    if not id_lines:
        message.reply_text("Tidak ada ID yang ditemukan setelah tanggal.")
        return

    # ------------------------------------------------------
    # 3. SIMPAN FOTO (sementara ke file)
    # ------------------------------------------------------
    file = message.photo[-1].get_file()
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    foto_temp_path = os.path.join(FOLDER_FOTO, f"temp_{timestamp}.jpg")
    file.download(foto_temp_path)

    # ------------------------------------------------------
    # 4. CEK FOTO DUPLIKAT (HASH)
    # ------------------------------------------------------
    foto_hash = generate_photo_hash(foto_temp_path)

    for anggota in data:
        if "History" in anggota:
            for h in anggota["History"]:
                if "Hash" in h and h["Hash"] == foto_hash:
                    # Foto ini pernah dipakai sebelumnya
                    message.reply_text(
                        "⚠ Foto ini sama persis dengan foto sebelumnya.\n"
                        "Kunjungan DITOLAK."
                    )
                    os.remove(foto_temp_path)
                    return

    # ------------------------------------------------------
    # 5. SIMPAN FOTO DENGAN NAMA FINAL
    # ------------------------------------------------------
    foto_path = os.path.join(FOLDER_FOTO, f"kunjungan_{timestamp}.jpg")
    os.rename(foto_temp_path, foto_path)

    # ------------------------------------------------------
    # 6. PROSES SEMUA BARIS ID
    # ------------------------------------------------------
    updated_count = 0
    errors = []

    for line in id_lines:
        id_text, payment = parse_id_payment_line(line)

        if id_text is None:
            errors.append(f"Format salah: {line}")
            continue

        if payment is None:
            errors.append(f"Nominal salah: {line}")
            continue

        found = False

        for d in data:
            if d["ID"] == id_text:
                found = True

                # Update status
                d["Status"] = "Sudah Dikunjungi"
                d["Foto"] = foto_path
                d["TanggalKunjungan"] = tanggal_kunjungan

                # Tambah HISTORY baru
                d["History"].append({
                    "Tanggal": tanggal_kunjungan,
                    "Payment": payment,
                    "Foto": foto_path,
                    "Hash": foto_hash
                })

                updated_count += 1
                break

        if not found:
            errors.append(f"ID tidak ditemukan: {id_text}")

    save_data(data)

    # ------------------------------------------------------
    # 7. BALASAN KE TELEGRAM
    # ------------------------------------------------------
    reply = (
        f"✔ {updated_count} anggota berhasil dicatat!\n"
        f"📅 Tanggal kunjungan: {tanggal_kunjungan}\n"
    )

    if errors:
        reply += "\n⚠ Baris dengan error:\n"
        for e in errors:
            reply += f"- {e}\n"

    message.reply_text(reply)
# ==========================================================
# BAGIAN 4 – PERINTAH /HAPUS, /HISTORY, /LIST_KUNJUNGAN
# ==========================================================

# ----------------------------------------------------------
# 1. /hapus ID – Menghapus kunjungan terakhir anggota
# ----------------------------------------------------------

def hapus_kunjungan(update: Update, context: CallbackContext):
    message = update.message
    parts = message.text.split()

    if len(parts) != 2:
        message.reply_text("Format:\n/hapus ID")
        return

    id_target = parts[1].strip()

    data = load_data()
    ensure_history_structure(data)

    found = False

    for d in data:
        if d["ID"] == id_target:
            found = True

            # Reset status
            d["Status"] = "Belum Dikunjungi"
            d["Foto"] = None
            d["TanggalKunjungan"] = None

            # History TIDAK DIHAPUS (sesuai permintaan)
            save_data(data)

            message.reply_text(
                f"✔ Kunjungan untuk ID {id_target} berhasil dibatalkan.\n"
                f"Status kembali: Belum Dikunjungi"
            )
            break

    if not found:
        message.reply_text(f"ID {id_target} tidak ditemukan di database.")


# ----------------------------------------------------------
# 2. /history ID – Menampilkan riwayat pembayaran & tanggal
# ----------------------------------------------------------

def show_history(update: Update, context: CallbackContext):
    message = update.message
    parts = message.text.split()

    if len(parts) != 2:
        message.reply_text("Format:\n/history ID")
        return

    id_target = parts[1].strip()

    data = load_data()
    ensure_history_structure(data)

    found = False

    for d in data:
        if d["ID"] == id_target:
            found = True
            nama = d["Nama"]

            if not d["History"]:
                message.reply_text(
                    f"📜 History Pembayaran: {nama}\n"
                    f"ID: {id_target}\n\n"
                    f"Belum ada riwayat pembayaran."
                )
                return

            text = (
                f"📜 History Pembayaran: {nama}\n"
                f"ID: {id_target}\n\n"
            )

            # Urutkan berdasarkan tanggal
            sorted_history = sorted(
                d["History"],
                key=lambda x: datetime.datetime.strptime(x["Tanggal"], "%d-%m-%Y")
            )

            for h in sorted_history:
                tanggal = h["Tanggal"]
                payment = h["Payment"]
                text += f"{tanggal} → Rp.{payment:,}\n"

            message.reply_text(text)
            return

    if not found:
        message.reply_text(f"ID {id_target} tidak ditemukan.")


# ----------------------------------------------------------
# 3. /list_kunjungan – Export Excel anggota BELUM DIKUNJUNGI
# ----------------------------------------------------------

def list_kunjungan(update: Update, context: CallbackContext):
    data = load_data()

    belum = [d for d in data if d["Status"] == "Belum Dikunjungi"]

    if not belum:
        update.message.reply_text("Semua anggota sudah dikunjungi. 🎉")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Belum Dikunjungi"

    headers = ["No", "Ctr", "ID", "Nama", "Minggon", "NamaStaff", "Status"]
    ws.append(headers)

    for d in belum:
        ws.append([
            d["No"],
            d["Ctr"],
            d["ID"],
            d["Nama"],
            d["Minggon"],
            d["NamaStaff"],
            d["Status"]
        ])

    file_name = f"Belum_Dikunjungi_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file_path = os.path.join(EXPORT_FOLDER, file_name)

    wb.save(file_path)

    with open(file_path, "rb") as f:
        update.message.reply_document(f, filename=file_name)

    os.remove(file_path)
# ==========================================================
# BAGIAN 5 – FITUR /REKAP + HELP + MAIN
# ==========================================================

# ----------------------------------------------------------
# 1. /rekap dd-mm-yyyy dd-mm-yyyy
# ----------------------------------------------------------

def rekap(update: Update, context: CallbackContext):
    msg = update.message.text.split()

    # ------------------------------------------------------
    # FORMAT 1 → /rekap tanggal_awal tanggal_akhir
    # FORMAT 2 → /rekap tanggal (rekap per hari)
    # ------------------------------------------------------

    if len(msg) == 2:
        # Format /rekap 01-12-2025
        tgl_awal = msg[1]
        tgl_akhir = msg[1]

    elif len(msg) == 3:
        # Format /rekap 01-12-2025 07-12-2025
        tgl_awal = msg[1]
        tgl_akhir = msg[2]

    else:
        update.message.reply_text(
            "Format salah.\nGunakan:\n"
            "/rekap 01-12-2025\n"
            "/rekap 01-12-2025 07-12-2025"
        )
        return

    # Validasi tanggal
    try:
        d_awal = datetime.datetime.strptime(tgl_awal, "%d-%m-%Y")
        d_akhir = datetime.datetime.strptime(tgl_akhir, "%d-%m-%Y")
    except:
        update.message.reply_text("Format tanggal harus dd-mm-yyyy.")
        return

    data = load_data()
    ensure_history_structure(data)

    # Struktur rekap
    staff_data = {}

    for d in data:
        staff = d["NamaStaff"]
        if staff not in staff_data:
            staff_data[staff] = {
                "ctr": set(),
                "agt": 0,
                "pay": 0
            }

        for h in d["History"]:
            tgl = datetime.datetime.strptime(h["Tanggal"], "%d-%m-%Y")

            if d_awal <= tgl <= d_akhir:
                staff_data[staff]["ctr"].add(d["Ctr"])
                staff_data[staff]["agt"] += 1
                staff_data[staff]["pay"] += h["Payment"]

    # Buat output
    output = f"📌 Report ALL {tgl_awal} s/d {tgl_akhir}\n\n"

    total_ctr = 0
    total_agt = 0
    total_pay = 0

    for staff, vals in staff_data.items():
        c = len(vals["ctr"])
        a = vals["agt"]
        p = vals["pay"]

        total_ctr += c
        total_agt += a
        total_pay += p

        output += (
            f"👤 {staff}\n"
            f"📍 total ctr : {c}\n"
            f"📍 total agt : {a}\n"
            f"💰 total pay : Rp.{p:,}\n\n"
        )

    output += (
        "================================\n"
        f"📍 t.ctr : {total_ctr} | "
        f"📍 t.agt : {total_agt} | "
        f"💰 t.pay : Rp.{total_pay:,}"
    )

    update.message.reply_text(output)



# ----------------------------------------------------------
# 2. /help – Daftar perintah
# ----------------------------------------------------------

def help_cmd(update: Update, context: CallbackContext):
    update.message.reply_text(
        "📌 DAFTAR PERINTAH BOT\n\n"
        "/import_excel – Upload file Excel anggota\n"
        "/list_kunjungan – Export Excel anggota belum dikunjungi\n"
        "/hapus ID – Batalkan kunjungan anggota\n"
        "/history ID – Lihat riwayat bayar anggota\n"
        "/rekap dd-mm-yyyy dd-mm-yyyy – Rekap laporan\n\n"
        "Kirim FOTO dengan format:\n"
        "dd-mm-yyyy (opsional)\n"
        "ID payment\n"
        "ID payment\n\n"
        "Jika tanpa tanggal → otomatis pakai tanggal hari ini."
    )


# ----------------------------------------------------------
# 3. MAIN() – MENJALANKAN BOT TELEGRAM
# ----------------------------------------------------------

def main():
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher

    # Handler
    dp.add_handler(CommandHandler("import_excel", import_excel))
    dp.add_handler(CommandHandler("list_kunjungan", list_kunjungan))
    dp.add_handler(CommandHandler("hapus", hapus_kunjungan))
    dp.add_handler(CommandHandler("history", show_history))
    dp.add_handler(CommandHandler("rekap", rekap))
    dp.add_handler(CommandHandler("help", help_cmd))

    # Foto kunjungan (tanpa command)
    dp.add_handler(MessageHandler(Filters.photo, foto_kunjungan))

    # Mulai bot
    updater.start_polling()
    updater.idle()


# ----------------------------------------------------------
# JALANKAN BOT
# ----------------------------------------------------------

if __name__ == "__main__":
    main()


